#!/usr/bin/env python3
"""Export the prediction workbook into static JSON for the web app."""

from __future__ import annotations

import argparse
import json
import warnings
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl

warnings.filterwarnings(
    "ignore",
    message="Conditional Formatting extension is not supported and will be removed",
    category=UserWarning,
)


PARTICIPANT_START_COL = 18
PARTICIPANT_COUNT = 33
MATCH_START_ROW = 8
MATCH_END_ROW = 155
MATCH_MAX_COL = 210


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def as_int(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def match_stage(match_id: int) -> str:
    if match_id <= 72:
        return "Group Round"
    if match_id <= 88:
        return "Round of 32"
    if match_id <= 96:
        return "Round of 16"
    if match_id <= 100:
        return "Quarterfinals"
    if match_id <= 102:
        return "Semi-Finals"
    if match_id == 103:
        return "Third-Place Play-Off"
    if match_id == 104:
        return "Final"
    return "Tournament"


def export_workbook(source: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
    if "2026 World Cup" not in workbook.sheetnames:
        raise SystemExit("Workbook is missing the '2026 World Cup' sheet.")

    sheet = workbook["2026 World Cup"]
    settings = workbook["Settings"] if "Settings" in workbook.sheetnames else None

    header = next(sheet.iter_rows(min_row=5, max_row=5, max_col=MATCH_MAX_COL, values_only=True))
    stage_prizes: dict[str, dict[str, int]] = {}
    for row in sheet.iter_rows(min_row=8, max_row=17, min_col=206, max_col=210, values_only=True):
        stage = clean(row[0])
        prize_per_person = as_int(row[1])
        prize_per_match = as_int(row[2])
        match_count = as_int(row[3])
        total_prize = as_int(row[4])
        if stage and prize_per_match is not None:
            stage_prizes[str(stage)] = {
                "prizePerPerson": prize_per_person or 0,
                "prizePerMatch": prize_per_match,
                "matchCount": match_count or 0,
                "totalPrize": total_prize or 0,
            }

    participants: list[str] = []
    participant_cols: list[int] = []
    for index in range(PARTICIPANT_COUNT):
        col = PARTICIPANT_START_COL + index * 2
        name = clean(header[col - 1] if col - 1 < len(header) else None)
        if name:
            participants.append(str(name))
            participant_cols.append(col)

    matches: list[dict[str, Any]] = []
    for row in sheet.iter_rows(
        min_row=MATCH_START_ROW,
        max_row=MATCH_END_ROW,
        max_col=MATCH_MAX_COL,
        values_only=True,
    ):
        match_id = as_int(row[1] if len(row) > 1 else None)
        if not match_id:
            continue

        predictions: dict[str, list[int]] = {}
        for name, col in zip(participants, participant_cols):
            home_guess = as_int(row[col - 1] if col - 1 < len(row) else None)
            away_guess = as_int(row[col] if col < len(row) else None)
            if home_guess is not None and away_guess is not None:
                predictions[name] = [home_guess, away_guess]

        actual_home = as_int(row[7] if len(row) > 7 else None)
        actual_away = as_int(row[8] if len(row) > 8 else None)
        seed_score = [actual_home, actual_away] if actual_home is not None and actual_away is not None else None
        stage = match_stage(match_id)
        prize_info = stage_prizes.get(stage, {})

        matches.append(
            {
                "id": match_id,
                "day": clean(row[2] if len(row) > 2 else None),
                "date": clean(row[3] if len(row) > 3 else None),
                "time": clean(row[4] if len(row) > 4 else None),
                "home": clean(row[5] if len(row) > 5 else None),
                "away": clean(row[10] if len(row) > 10 else None),
                "group": clean(row[11] if len(row) > 11 else None),
                "venue": clean(row[12] if len(row) > 12 else None),
                "stage": stage,
                "prizePerMatch": prize_info.get("prizePerMatch", 0),
                "prizePerPerson": prize_info.get("prizePerPerson", 0),
                "seedScore": seed_score,
                "status": "SCHEDULED",
                "predictions": predictions,
            }
        )

    timezone = None
    minute_offset = None
    if settings:
        timezone = clean(settings["C8"].value)
        minute_offset = clean(settings["C10"].value)

    return {
        "generatedAt": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "sourceWorkbook": str(source),
        "settings": {
            "title": "2026 World Cup Prediction Pool",
            "moneyLabel": "Prize",
            "moneySuffix": "",
            "refreshSeconds": 60,
            "timezone": timezone,
            "minuteOffset": minute_offset,
            "stagePrizes": stage_prizes,
        },
        "participants": participants,
        "matches": sorted(matches, key=lambda item: item["id"]),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="Path to the 2026 FIFA World Cup workbook")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("public/data/predictions.json"),
        help="JSON file to write",
    )
    args = parser.parse_args()

    payload = export_workbook(args.source.expanduser().resolve())
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(
        f"Exported {len(payload['matches'])} matches and "
        f"{len(payload['participants'])} participants to {args.output}"
    )


if __name__ == "__main__":
    main()
